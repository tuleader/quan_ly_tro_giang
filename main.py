import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from tkcalendar import Calendar, DateEntry
import json
import os
from datetime import datetime
import openpyxl
# Cấu hình cửa sổ chính
root = tk.Tk()
root.title("Quản lý trợ giảng")
root.geometry("360x780")
# Khởi tạo biến
file_du_lieu = "du_lieu_nhan_vien.json"
if os.path.exists(file_du_lieu):
    with open(file_du_lieu, "r") as file:
        nhan_vien = json.load(file)
else:
    nhan_vien = {}

# Hàm lưu dữ liệu vào file
def luu_du_lieu():
    with open(file_du_lieu, "w") as file:
        json.dump(nhan_vien, file)

# Hàm thêm nhân viên
def them_nhan_vien():
    ten = simpledialog.askstring("Nhập Tên", "Tên nhân viên:")
    if ten and ten not in nhan_vien:
        nhan_vien[ten] = {}
        luu_du_lieu()
        cap_nhat_danh_sach()
    elif ten:
        messagebox.showinfo("Thông báo", "Nhân viên đã tồn tại.")

# Hàm xóa nhân viên
def xoa_nhan_vien():
    ten = danh_sach.get(tk.ANCHOR)
    if ten and messagebox.askyesno("Xác nhận", f"Bạn có chắc chắn muốn xóa {ten}?"):
        del nhan_vien[ten]
        luu_du_lieu()
        cap_nhat_danh_sach()

# Hàm cập nhật danh sách nhân viên cho ngày được chọn
def cap_nhat_danh_sach_cham_cong(ngay):
    danh_sach.delete(0, tk.END)
    for ten, cham_cong in nhan_vien.items():
        gio_lam = cham_cong.get(ngay, 0)
        danh_sach.insert(tk.END, f"{ten} - {gio_lam} giờ")

# Hàm cập nhật giờ làm việc khi nhấn vào danh sách
def cap_nhat_gio_lam(event):
    chon = danh_sach.curselection()
    if chon:
        ten, gio = danh_sach.get(chon[0]).split(' - ')
        gio_hien_tai = float(gio.split(' ')[0])
        gio_moi = simpledialog.askinteger("Cập nhật giờ làm", f"Nhập số giờ làm cho {ten}:", initialvalue=gio_hien_tai)
        if gio_moi is not None:
            ngay = cal.get_date()
            nhan_vien[ten][ngay] = gio_moi
            luu_du_lieu()
            cap_nhat_danh_sach_cham_cong(ngay)

# Hàm xử lý thay đổi ngày trên lịch
def on_ngay_thay_doi(event):
    ngay_chon = cal.get_date()
    cap_nhat_danh_sach_cham_cong(ngay_chon)

# Hàm thống kê giờ làm trong tháng
def thong_ke_thang():
    thang = thang_thong_ke.get_date().strftime("%m/%Y")
    tong_gio_thang = {}
    for ten, cham_cong in nhan_vien.items():
        tong_gio = sum(gio for ngay, gio in cham_cong.items() if ngay.endswith(thang))
        tong_gio_thang[ten] = tong_gio
    hien_thi_thong_ke(thang,tong_gio_thang)

def hien_thi_thong_ke(thang, tong_gio_thang):
    ket_qua_thong_ke = "\n".join(f"{ten}: {gio} giờ" for ten, gio in tong_gio_thang.items())
    
    # Hiển thị hộp thoại và lấy kết quả
    user_response = messagebox.askquestion(f"Thống Kê Tháng {thang}", ket_qua_thong_ke)
    
    # Nếu người dùng chọn "OK", thì xuất file Excel
    if user_response == 'yes':
        # Tạo một file Excel và ghi kết quả thống kê vào đó
        workbook = openpyxl.Workbook()
        
        # Thay thế các ký tự không hợp lệ trong tên sheet
        thang_moi = thang.replace('/', '_')
        sheet_title = f"Thống_Kê_Tháng {thang_moi}"
        sheet = workbook.active
        sheet.title = sheet_title

        # Ghi dữ liệu vào Excel
        for row, (ten, gio) in enumerate(tong_gio_thang.items(), start=1):
            sheet.cell(row=row, column=1, value=ten)
            sheet.cell(row=row, column=2, value=gio)

        # Tạo thư mục để chứa file nếu nó chưa tồn tại
        os.makedirs("thong_ke_cham_cong", exist_ok=True)
        
        # Lưu file Excel
        excel_filename = os.path.join("thong_ke_cham_cong", f"thong_ke_thang_{thang_moi}.xlsx")
        workbook.save(excel_filename)
        messagebox.showinfo("Xuất Excel", f"Kết quả thống kê đã được xuất thành công vào file: {excel_filename}")

# Hàm cập nhật danh sách nhân viên
def cap_nhat_danh_sach():
    danh_sach.delete(0, tk.END)
    for ten in nhan_vien.keys():
        danh_sach.insert(tk.END, ten)
# Hàm hiển thị danh sách nhân viên
def hien_thi_danh_sach_nhan_vien():
    cap_nhat_danh_sach()

# Giao diện
tk.Button(root, text="Hiển Thị Danh Sách Trợ Giảng", command=hien_thi_danh_sach_nhan_vien).place(x=80, y=10, width=200, height=30)
tk.Button(root, text="Thêm Trợ Giảng", command=them_nhan_vien).place(x=130, y=50, width=100, height=30)
tk.Button(root, text="Xóa Trợ Giảng", command=xoa_nhan_vien).place(x=130, y=90, width=100, height=30)

danh_sach = tk.Listbox(root)
danh_sach.place(x=30, y=130, width=300, height=250)
cap_nhat_danh_sach()

cal = Calendar(root, selectmode='day', date_pattern='dd/mm/yyyy')
cal.place(x=30, y=390, width=300, height=150)
cal.bind("<<CalendarSelected>>", on_ngay_thay_doi)
danh_sach.bind('<Double-1>', cap_nhat_gio_lam)
thang_thong_ke = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2, year=datetime.now().year, month=datetime.now().month, date_pattern='dd/mm/yyyy')
thang_thong_ke.place(x=30, y=550, width=300, height=30)

tk.Button(root, text="Thống Kê Tháng", command=thong_ke_thang).place(x=130, y=590, width=100, height=30)
root.mainloop()